import csv
import datetime
import os
from csv import DictReader
from select import error
from venv import create

import django_filters
from dateutil.relativedelta import relativedelta
from django.core.mail import EmailMessage
from django.forms import DateInput
from django.http import HttpResponseRedirect
from django.utils.translation import gettext_lazy as _
from django.contrib.auth import logout
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth.models import User
from django.contrib.auth.views import LoginView, PasswordChangeView
from django.shortcuts import render, get_object_or_404, redirect
from datetime import timedelta
from django import forms
from django.db.models import Q
from django.utils.timezone import now
from django.urls import reverse_lazy
from django.views.generic import UpdateView, CreateView, ListView, DetailView, DeleteView
from django_extensions.templatetags.widont import widont
from django_filters.filters import _truncate
from django_filters.views import FilterView
from openpyxl.reader.excel import load_workbook
from rest_framework.permissions import IsAdminUser, IsAuthenticated
from urllib3 import request

from device.forms import AddEquipmentForm, AddDeviceForm, DraftForm, LoginUserForm, MyExamsForm, PprForm
from device.models import Equipment, GP, Si, EquipmentModel, Manufacturer, Status, Position, \
    EquipmentName, Location, Tag, StatusAdd, Description, Year, Draft, VerificationInterval, Unit, Scale, \
    MyExam, PprDate, PprPlan, Manual, Category
# MyExam
# from device.parser import data_from_parser
from device.sending import sample_send
from device.variables import year
from equipment.settings import BASE_DIR
from django.http import HttpResponseRedirect
from django.shortcuts import render
from django.core.files import File
menu = [
    {'title': 'Модели', 'url_name': 'models'},
    {'title': 'Производители', 'url_name': 'manufacturers'},
    {'title': 'Названия', 'url_name': 'names'},
    {'title': 'Статусы', 'url_name': 'statuses'},
    {'title': 'Года выпуска', 'url_name': 'years'},
    {'title': 'Позиции', 'url_name': 'list_gp'},
    {'title': 'Поиск', 'url_name': 'search'},
    {'title': 'Подписанты', 'url_name': 'defectone:approves'},
    {'title': 'Мастера', 'url_name': 'defectone:kaits'},
    {'title': 'Мастера по цеху', 'url_name': 'defectone:workers'},
    {'title': 'Подрядчики', 'url_name': 'defectone:contractors'},
    {'title': 'Дефекты', 'url_name': 'defectone:defect_list'}
]



def si_loading(request, i):
    u = User.objects.get(username='admin')
    StatusAdd.objects.get_or_create(name='Установлен')
    status = StatusAdd.objects.get(name='Установлен')
    Manufacturer.objects.get_or_create(name='manufacturer')
    man = Manufacturer.objects.get(name='manufacturer')
    with open(f'./si{i}.csv', encoding='utf-8') as si:
        reader = csv.DictReader(si, delimiter=';')
        # GP.objects.all().delete()
        for row in reader:
            if Equipment.objects.filter(Q(serial_number=row['serial_number']) & Q(model__name=row['model'])):
                continue
            rez = row['previous_verification'].strip().split('.')
            str_list = [i for i in rez if i]
            d = '{}-{}-{}'.format(str_list[2], str_list[1], str_list[0])
            previous_verification = datetime.date.fromisoformat(d)
            next_verification = previous_verification + relativedelta(
                months=+(int(row['interval'])))
            EquipmentModel.objects.get_or_create(name=row['model'].strip().capitalize())
            m = EquipmentModel.objects.get(name=row['model'].strip().capitalize())
            # EquipmentType.objects.get_or_create(name=row['type'].strip().capitalize())
            # t = EquipmentType.objects.get(name=row['type'].strip().capitalize())
            EquipmentName.objects.get_or_create(name=row['name'].strip().capitalize())
            n = EquipmentName.objects.get(name=row['name'].strip().capitalize())
            Year.objects.get_or_create(name=row['year'])
            y = Year.objects.get(name=row['year'])
            #     GP.objects.get_or_create(name=row['position'].upper(), construction=row['construction'])
            #     gp = GP.objects.get(name=row['position'])
            #     gp.construction = row['construction']
            #     gp.save()
            try:
                eq = Equipment.objects.create(
                    serial_number=row['serial_number'],
                    model=m,
                    si_or=True,
                    manufacturer=man,
                    # type=t,
                    name=n,
                    year=y,
                )
            except:
                continue

            Status.objects.create(name=status, equipment=eq)
            Position.objects.create(name=row['position'].strip().upper(), equipment=eq)
            Location.objects.create(name=row['location'].strip().capitalize(), equipment=eq)
            Tag.objects.create(name=row['teg'].capitalize(), equipment=eq)
            Description.objects.create(name='description'.capitalize(), equipment=eq, user=u)
            VerificationInterval.objects.get_or_create(name=row['interval'])
            interval = VerificationInterval.objects.get(name=row['interval'])
            Scale.objects.get_or_create(min_scale=row['min_scale'], max_scale=row['max_scale'])
            scale = Scale.objects.get(min_scale=row['min_scale'], max_scale=row['max_scale'])
            Unit.objects.get_or_create(name=row['unit'])
            unit = Unit.objects.get(name=row['unit'])
            # RegNumber.objects.get_or_create(name=row['reg_number'])
            # reg_number = RegNumber.objects.get(name=row['reg_number'])
            if row['result'] == 'Годен':
                rezult = True
            else:
                rezult = False
            Si.objects.create(
                equipment=eq,
                previous_verification=previous_verification,
                next_verification=next_verification,
                # certificate=row['certificate'],
                interval=interval,
                scale=scale,
                unit=unit,
                # reg_number=reg_number,
                result=rezult,
                com=row['comment']
            )

        return render(request, 'device/equipments.html')


def IM(request):
    u = User.objects.get(username='admin')
    with open('./im.csv') as f:
        readers = csv.DictReader(f, delimiter=';')
        for i, row in enumerate(readers):
            EquipmentModel.objects.get_or_create(name=row['model'])
            m = EquipmentModel.objects.get(name=row['model'])
            # EquipmentType.objects.get_or_create(name=row['type'])
            # t = EquipmentType.objects.get(name=row['type'])
            EquipmentName.objects.get_or_create(name=row['name'].capitalize())
            n = EquipmentName.objects.get(name=row['name'])
            if (row['type'] == 'РэмТэк') and (len(row['serial_number']) > 4):
                year_eq = year[(row['serial_number'][0:2])]
            else:
                year_eq = '1999'
            Year.objects.get_or_create(name=year_eq)
            y = Year.objects.get(name=year_eq)
            Manufacturer.objects.get_or_create(name='manufacturer')
            man = Manufacturer.objects.get(name='manufacturer')
            try:
                eq = Equipment.objects.create(
                    serial_number=row['serial_number'],
                    model=m,
                    si_or=False,
                    manufacturer=man,
                    # type=t,
                    name=n,
                    year=y,
                )
            except:
                continue
            Position.objects.create(name=row['poz'].upper(), equipment=eq)
            Location.objects.create(name=row['location'].capitalize(), equipment=eq)
            Tag.objects.create(name=row['tag'], equipment=eq)
            Description.objects.create(name='description'.capitalize(), equipment=eq, user=u)
            status = StatusAdd.objects.get(name='Установлен')
            Status.objects.create(name=status, equipment=eq)

        return render(request, 'device/equipments.html')


def equipment_add(request):
    if not request.user.is_staff:
        redirect('login')
    if request.method == 'POST':
        form = AddEquipmentForm(request.POST)
        if form.is_valid():
            form.save(request.user)
            return redirect('/')
    else:
        form = AddEquipmentForm()
    models = EquipmentModel.objects.all()

    return render(request, 'device/equipment_add.html', {'form': form, 'menu': menu, 'models': models})


def device_add(request):
    if not request.user.is_staff:
        redirect('login')
    if request.method == 'POST':
        form = AddDeviceForm(request.POST)
        if form.is_valid():
            form.save(request.user)
            return redirect('/')
    else:
        form = AddDeviceForm()
    return render(request, 'device/equipment_add.html', {'form': form, 'menu': menu})


def EquipmentUpdate(request, pk):
    if not request.user.is_authenticated:
        redirect('login')
    equipment = get_object_or_404(Equipment, pk=pk)
    status = StatusAdd.objects.all()
    positions = GP.objects.all()
    data = {
        'equipment': equipment,
        'menu': menu,
        'status': status,
        'positions': positions,
        'si_or': False,
    }
    if request.method == 'POST':
        # if request.POST['description'] != equipment.descriptions.last().name:
        Tag.objects.create(equipment=equipment, name=request.POST['tag'])
        if len(request.POST['location']) > 255:
            data['error'] = 'Местоположение не может содержать более 255 символов'
            return render(request, 'device/equipment_update.html', context=data)
        try:
            if request.POST['defect_or']:
                equipment.defect_or = True
                equipment.save()
        except:
            pass
        if request.POST.get('position_new'):
            poz = request.POST.get('position_new')
            GP.objects.get_or_create(name=poz)
        else:
            poz = request.POST['position']
        Location.objects.create(equipment=equipment, name=request.POST['location'])
        Position.objects.create(equipment=equipment, name=poz.upper())
        Description.objects.create(equipment=equipment, user=request.user, name=request.POST['description'])
        status = StatusAdd.objects.get(name=request.POST['status'])
        Status.objects.create(equipment=equipment, name=status)
        return redirect('search')
    # else:
    #     data['error'] = 'Комментарий не был изменен'

    return render(request, 'device/equipment_update.html', context=data)


class MyFilter(django_filters.FilterSet):
    # type = django_filters.CharFilter(field_name='type__name',
    #                                  lookup_expr='icontains',
    #                                  label='Тип:',
    #                                  widget=forms.TextInput(attrs={'class': 'type2'}))
    serial_number = django_filters.CharFilter(lookup_expr='icontains',
                                              widget=forms.TextInput(attrs={'class': 'type2'}),
                                              label='Серийный номер')
    # model = django_filters.CharFilter(field_name='model__name',
    #                                   lookup_expr='icontains',
    #                                   label='Модель:',
    #                                   widget=forms.TextInput(attrs={'class': 'type2'}))
    position = django_filters.ModelChoiceFilter(widget=forms.Select(attrs={'class': 'select'}),
                                                queryset=GP.objects.all(),
                                                field_name='positions__name',
                                                lookup_expr='exact', label='Позиция:', )
    locations = django_filters.CharFilter(widget=forms.TextInput(attrs={'class': 'type2'}),
                                          field_name='locations__name',
                                          lookup_expr='icontains', label='Место установки:', )
    tag = django_filters.CharFilter(widget=forms.TextInput(attrs={'class': 'type2'}),
                                    field_name='tags__name',
                                    lookup_expr='icontains', label='Тэг:', )
    name = django_filters.CharFilter(field_name='name__name', lookup_expr='icontains', label='Наименование:',
                                     widget=forms.TextInput(attrs={'class': 'type2'}))

    start_date = django_filters.DateFilter(
        widget=forms.TextInput(attrs=
        {
            'type': 'date',
            'class': 'type2',
        }), label='Поверка от:',
        field_name='si__next_verification', lookup_expr='gte', )
    end_date = django_filters.DateFilter(
        widget=forms.TextInput(attrs=
        {
            'type': 'date',
            'class': 'type2',
        }), label='Поверка до:',
        field_name='si__next_verification', lookup_expr='lte')

    start_date_add = django_filters.DateFilter(
        widget=forms.TextInput(attrs=
        {
            'type': 'date',
            'class': 'type2',
        }), label='Внесено с:',
        field_name='descriptions__at_date', lookup_expr='gte', )
    end_date_add = django_filters.DateFilter(
        widget=forms.TextInput(attrs=
        {
            'type': 'date',
            'class': 'type2',
        }), label='Внесено по:',
        field_name='descriptions__at_date', lookup_expr='lte')

    # choices = [
    #     ("today", _("Today")),
    #     ("week", _("Past 7 days")),
    #     ("month", _("This month")),
    # ]
    # filters = {
    #     "today": lambda qs, name: qs.filter(
    #         **{
    #             "%s__year" % name: now().year,
    #             "%s__month" % name: now().month,
    #             "%s__day" % name: now().day,
    #         }
    #     ),
    #
    #     "week": lambda qs, name: qs.filter(
    #         **{
    #             "%s__gte" % name: _truncate(now() - timedelta(days=7)),
    #             "%s__lt" % name: _truncate(now() + timedelta(days=1)),
    #         }
    #     ),
    #     "month": lambda qs, name: qs.filter(
    #         **{"%s__year" % name: now().year, "%s__month" % name: now().month}
    #     ),
    # }

    # date_range = django_filters.DateRangeFilter(
    #     widget=forms.Select(attrs=
    #     {
    #         'class': 'type2',
    #     }),
    #     label='Поверка:', filters=filters, choices=choices, field_name='si__next_verification')

    # at_date = django_filters.DateFilter(
    #     widget=forms.TextInput(attrs=
    #     {
    #         'type': 'date',
    #         'class': 'type2',
    #     }), label='Внесено в этот день',
    #     field_name='descriptions__at_date', lookup_expr='contains')
    status = django_filters.ModelChoiceFilter(widget=forms.Select(attrs={'class': 'select'}),
                                              queryset=StatusAdd.objects.all(), field_name='status__name',
                                              lookup_expr='exact', label='Статус')
    model = django_filters.CharFilter(widget=forms.TextInput(attrs={'class': 'type2'}),
                                      field_name='model__name',
                                      lookup_expr='icontains', label='Модель')
    manufacturer = django_filters.ModelChoiceFilter(widget=forms.Select(attrs={'class': 'select'}),
                                                    queryset=Manufacturer.objects.all(),
                                                    field_name='manufacturer__name',
                                                    lookup_expr='exact', label='Производитель')
    si_or = django_filters.BooleanFilter(field_name='si_or', widget=forms.NullBooleanSelect(attrs={'class': 'select'}))
    defect_or = django_filters.BooleanFilter(field_name='defect_or',
                                             widget=forms.NullBooleanSelect(attrs={'class': 'select'}))

    class Meta:
        model = Equipment
        fields = ['serial_number', 'name', 'model', 'position', 'locations', 'tag', 'status', 'si_or',
                  'manufacturer', 'defect_or', ]


class MyFilterUser(django_filters.FilterSet):
    # type = django_filters.CharFilter(field_name='type__name',
    #                                  lookup_expr='icontains',
    #                                  label='Тип:',
    #                                  widget=forms.TextInput(attrs={'class': 'type2'}))
    serial_number = django_filters.CharFilter(lookup_expr='icontains',
                                              widget=forms.TextInput(attrs={'class': 'type2'}),
                                              label='Серийный номер')
    position = django_filters.ModelChoiceFilter(widget=forms.Select(attrs={'class': 'select'}),
                                                queryset=GP.objects.all(),
                                                field_name='positions__name',
                                                lookup_expr='exact', label='Позиция:', )
    name = django_filters.CharFilter(field_name='name__name', lookup_expr='icontains', label='Наименование:',
                                     widget=forms.TextInput(attrs={'class': 'type2'}))
    status = django_filters.ModelChoiceFilter(widget=forms.Select(attrs={'class': 'select'}),
                                              queryset=StatusAdd.objects.all(), field_name='status__name',
                                              lookup_expr='exact', label='Статус')
    si_or = django_filters.BooleanFilter(widget=forms.NullBooleanSelect(attrs={'class': 'select'}))
    tag = django_filters.CharFilter(widget=forms.TextInput(attrs={'class': 'type2'}),
                                    field_name='tags__name',
                                    lookup_expr='icontains', label='Тэг:', )

    class Meta:
        model = Equipment
        fields = ['serial_number', 'name', 'position', 'si_or', 'status', 'tag']


def equipment_list(request):
    if not request.user.is_authenticated:
        redirect('/')
    if request.method == 'POST' and request.user.is_staff:
        eq_filter = MyFilter(request.POST,
                             queryset=Equipment.objects.prefetch_related('si', 'status', 'descriptions',
                                                                         'tags', ).all().order_by(
                                 'name'))
        s = set(eq_filter.qs)
        error_user = False
        error_staff = False
        try:
            object = MyExam.objects.get(user=request.user)
            objects = MyExam.objects.all()
            now_date = datetime.date.today() - relativedelta(months=14)
            if (object.exams_eb < now_date) or (object.exams_ot < now_date) and (not request.user.is_staff):
                error_user = True
            for obj in objects:
                if (obj.exams_eb < now_date) or (obj.exams_ot < now_date) and request.user.is_staff:
                    error_staff = True
        except:
            pass
        data = {
            'error_user': error_user,
            'error_staff': error_staff,
            'title': 'Поиск',
            'menu': menu,
            'equipments': s,
            'count': len(s),
            'forms': eq_filter,

        }
        sample_send(request, s)
        # man = Manufacturer.objects.get(name='ГазоАналит')
        # for i in eq_filter.qs:
        #     if 'СГОЭС' in i.model.name:
        #         i.manufacturer = man
        #         i.save()
        return render(request, 'device/equipments.html', context=data)
    if request.method == 'POST' and not request.user.is_staff:
        eq_filter = MyFilterUser(request.POST,
                                 queryset=Equipment.objects.prefetch_related('si', 'status', 'descriptions',
                                                                             'tags').all().order_by('name'))
        data = {
            'title': 'Поиск',
            'menu': menu,
            'equipments': eq_filter.qs,
            'count': eq_filter.qs.count(),
            'forms': eq_filter,

        }
        return render(request, 'device/equipments.html', context=data)

    if request.method == 'GET' and request.user.is_staff:
        eq_filter = MyFilter(request.POST,
                             queryset=Equipment.objects.all()[0:0])
    if request.method == 'GET' and not request.user.is_staff:
        eq_filter = MyFilterUser(request.POST,
                                 queryset=Equipment.objects.all()[0:0])
    data = {
        'title': 'Поиск',
        'menu': menu,
        'equipments': set(eq_filter.qs),
        'count': eq_filter.qs.count(),
        'forms': eq_filter,
    }
    return render(request, 'device/equipments.html', context=data)


def equipment_detail(request, pk):
    if not request.user.is_staff:
        redirect('/')
    equipment = get_object_or_404(Equipment, pk=pk)
    tag = equipment.tags.all()
    status = equipment.status.all()
    description = equipment.descriptions.all()
    location = equipment.locations.all()
    position = equipment.positions.all()
    statuses = []
    tags = []
    descriptions = []
    locations = []
    positions = []
    users = []
    data_eq = []
    at_date = []

    for i in position:
        positions.append(i.name)
    for i in status:
        statuses.append(i.name.name)
    for i in tag:
        tags.append(i.name)
    for i in description:
        descriptions.append(i.name)
        users.append(i.user)
        at_date.append(i.at_date)
    for i in location:
        locations.append(i.name)
    for number, i in enumerate(range(len(descriptions))):
        try:
            stat = statuses[i]
        except:
            stat = 'Нет статуса'

        data_eq.append({
            'description': descriptions[i],
            'tag': tags[i],
            'location': locations[i],
            'status': stat,
            'position': positions[i],
            'user': users[i],
            'at_date': at_date[i],
            'number': number + 1,
        })
        data_eq.reverse()
    data = {
        'equipment': equipment,
        'title': 'Информация об оборудовании',
        'menu': menu,
        'data_eq': data_eq,
    }
    return render(request, 'device/equipment_detail.html', context=data)


def DeviceUpdate(request, pk):
    if not request.user.is_authenticated:
        return redirect('login')
    equipment = get_object_or_404(Equipment, pk=pk)
    manufacturers = Manufacturer.objects.all()
    si = Si.objects.get(equipment=equipment)
    status = StatusAdd.objects.all()
    last_status = equipment.status.last()
    positions = GP.objects.all()
    last_position = equipment.positions.last()
    location = equipment.locations.last()
    description = equipment.descriptions.last()
    tag = equipment.tags.last()
    units = Unit.objects.all()
    min_scale = si.scale.min_scale
    max_scale = si.scale.max_scale
    unit = si.unit.name
    data = {
        'equipment': equipment,
        'manufacturers': manufacturers,
        'menu': menu,
        'status': status,
        'last_status': last_status,
        'positions': positions,
        'last_position': last_position,
        'location': location,
        'description': description,
        'tag': tag,
        'si_or': True,
        'units': units,
        'min_scale': min_scale,
        'max_scale': max_scale,
        'unit': unit,
    }
    if request.method == 'POST':
        # if request.POST['description'] != equipment.descriptions.last().name:
        t = Tag(equipment=equipment, name=request.POST['tag'])
        if len(request.POST['location']) > 255:
            data['error'] = 'Местоположение не может содержать более 255 символов'
            return render(request, 'device/equipment_update.html', context=data)
        l = Location(equipment=equipment, name=request.POST['location'])
        if request.POST.get('position_new'):
            poz = request.POST.get('position_new')
            GP.objects.get_or_create(name=poz)
        else:
            poz = request.POST['position']
        equipment.comment = request.POST['comment']
        if request.user.is_staff:
            equipment.manufacturer = Manufacturer.objects.get(name=request.POST['manufacturer'])
        try:
            request.POST['defect_or']
            equipment.defect_or = True
            equipment.save()
        except:
            equipment.defect_or = False
            equipment.save()
        p = Position(equipment=equipment, name=poz.upper())
        d = Description(equipment=equipment, user=request.user, name=request.POST['description'])
        status = StatusAdd.objects.get(name=request.POST['status'])
        s = Status(equipment=equipment, name=status)
        si.unit = Unit.objects.get(name=request.POST['unit'])
        Scale.objects.get_or_create(min_scale=request.POST['min_scale'], max_scale=request.POST['max_scale'])
        scale = Scale.objects.get(min_scale=request.POST['min_scale'], max_scale=request.POST['max_scale'])
        si.scale = scale
        si.scale.max_scale = '10'
        t.save()
        l.save()
        p.save()
        d.save()
        s.save()
        if not request.POST['previous_verification']:
            request.POST['previous_verification'] = '1990-01-01'
        si.previous_verification = request.POST['previous_verification']
        si.next_verification = (
                                   datetime.datetime.strptime(request.POST['previous_verification'],
                                                              '%Y-%m-%d').date()) + relativedelta(
            months=+int(si.interval.name))
        # try:
        #     si.certificate = request.POST['certificate']
        # except:
        #     si.certificate = '999999999'
        si.save()
        return redirect('search')
        # else:
        #     data['error'] = 'Комментарий не был изменен'
    return render(request, 'device/equipment_update.html', context=data)


def EquipmentDelete(request, pk):
    if not request.user.is_staff:
        redirect('login')
    if request.method == 'GET':
        data = {
            'equipment': get_object_or_404(Equipment, pk=pk),
            'menu': menu,
        }
        return render(request, 'device/equipment_delete.html', context=data)
    obj = get_object_or_404(Equipment, pk=pk)
    for d in obj.descriptions.all():
        d.delete()
    for p in obj.positions.all():
        p.delete()
    for l in obj.locations.all():
        l.delete()
    for t in obj.tags.all():
        t.delete()
    for s in obj.status.all():
        s.delete()
    for i in obj.si.all():
        i.delete()
    obj.delete()
    return redirect('/')


class MyFilterGp(django_filters.FilterSet):
    name = django_filters.CharFilter(lookup_expr='icontains')

    class Meta:
        model = GP
        fields = ['construction']


class ListGP(FilterView):
    model = GP
    permission_classes = [IsAdminUser, ]
    filterset_class = MyFilterGp
    template_name = 'device/gp_list.html'
    context_object_name = 'objects'
    extra_context = {
        'menu': menu
    }


class MyFilterModel(django_filters.FilterSet):
    name = django_filters.CharFilter(lookup_expr='icontains')

    class Meta:
        model = EquipmentModel
        fields = ['name']


class ListModel(FilterView):
    model = EquipmentModel
    permission_classes = [IsAdminUser, ]
    filterset_class = MyFilterModel
    template_name = 'device/list_model.html'
    context_object_name = 'objects'
    extra_context = {
        'menu': menu
    }


class AddGp(CreateView):
    model = GP
    permission_classes = [IsAdminUser, ]
    template_name = 'device/gp_add.html'
    fields = '__all__'
    extra_context = {
        'menu': menu
    }
    success_url = '/'


class UpdateGp(UpdateView):
    model = GP
    permission_classes = [IsAdminUser, ]
    template_name = 'device/gp_add.html'
    fields = '__all__'
    success_url = '/'
    extra_context = {
        'menu': menu,
    }


def delete_gp(request, pk):
    obj = get_object_or_404(GP, pk=pk)
    obj.delete()
    return redirect('list_gp')


class AddCategory(CreateView):
    permission_classes = [IsAdminUser, ]
    template_name = 'device/add_category.html'
    fields = ['name']
    context_object_name = 'cats'
    extra_context = {
        'menu': menu
    }
    success_url = '/'


class ListCategory(FilterView):
    permission_classes = [IsAdminUser, ]
    template_name = 'device/list_category.html'
    context_object_name = 'objects'


class UpdateCategory(UpdateView):
    permission_classes = [IsAdminUser, ]
    template_name = 'device/add_category.html'
    fields = ['name']
    context_object_name = 'cats'
    success_url = reverse_lazy('models')
    extra_context = {
        'menu': menu
    }

    def post(self, *args, **kwargs):
        self.object = self.get_object()
        # with open('./im.csv', encoding='utf-8') as f:
        #     reader = csv.DictReader(f, delimiter=';')
        #     rem = Equipment.objects.filter(type__name='РэмТэк')
        #     for row in reader:
        #         rem1 = rem.filter(serial_number=row['serial_number'])
        #         EquipmentModel.objects.get_or_create(name=f'РэмТэк {row["model"]}')
        #         m2 = EquipmentModel.objects.get(name=f'РэмТэк {row["model"]}')
        #         for j in rem1:
        #             j.model = m2
        #             j.save()
        # owner_model = EquipmentModel.objects.get(name=self.object.name)
        # name = EquipmentModel.objects.filter(name__icontains=self.object.name)
        #
        # count = 0
        # for i in name:
        #     count += 1
        #     if count == 1:
        #         continue
        #     eq = Equipment.objects.filter(model=i)
        #     for j in eq:
        #         try:
        #             j.model = owner_model
        #             j.save()
        #         except:
        #             pass
        #     try:
        #         i.delete()
        #     except:
        #         pass

        form = self.get_form()
        if form.is_valid():
            self.object = form.save(commit=False)
            self.object.save()
            return HttpResponseRedirect(self.get_success_url())
        else:
            return self.render_to_response(self.get_context_data(form=form))


# def delete_category(request, pk, Mod):
#     if not request.user.is_staff:
#         redirect('login')
#     obj = get_object_or_404(Mod, pk=pk)
#
#     obj.delete()
#     return redirect(reverse_lazy('search'))


class LoginUser(LoginView):
    template_name = 'users/login.html'
    extra_context = {
        'title': 'Авторизация',
        'menu': menu,
    }
    form_class = LoginUserForm
    success_url = '/'


class ChangePassword(PasswordChangeView):
    template_name = 'users/change_password.html'
    extra_context = {
        'title': 'Изменение пароля',
        'menu': menu,
    }
    form_class = PasswordChangeForm
    success_url = '/'


def logout_user(request):
    logout(request)
    return redirect('users:login')


class DraftCreate(CreateView):
    permission_classes = [IsAuthenticated, ]
    form_class = DraftForm
    template_name = 'device/draft.html'
    extra_context = {
        'title': 'Список черновиков',
        'menu': menu
    }

    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.user_draft = self.request.user
        instance.save()
        return redirect('/')


class DraftList(ListView):
    permission_classes = [IsAdminUser, ]
    model = Draft
    template_name = 'device/draft_list.html'
    context_object_name = 'drafts'
    queryset = Draft.objects.all()

    extra_context = {
        'title': 'Список черновиков',
        'menu': menu
    }


class DraftDetail(DetailView):
    permission_classes = [IsAdminUser, ]
    model = Draft
    template_name = 'device/draft_detail.html'


def draft_device_add(request, pk):
    if not request.user.is_staff:
        redirect('login')
    draft = Draft.objects.get(pk=pk)
    initial_dict = {
        'serial_number': draft.serial_number_draft,
        'position': draft.poz_draft,
        'model': draft.model_draft,
        'name': draft.name_draft,
        'manufacturer': draft.manufacturer_draft,
        'location': draft.location_draft,
        'tag': draft.tag_draft,
        'description': draft.description_draft,
        'status': draft.status_draft,
        'min_scale': draft.min_scale_draft,
        'max_scale': draft.max_scale_draft,
        'year': draft.year_draft,
        'unit': draft.unit_draft,

    }
    if request.method == 'POST':
        form = AddDeviceForm(request.POST)
        if form.is_valid():
            form.save(request.user)
            return redirect(reverse_lazy('draft_delete', kwargs={'pk': pk}))
    else:
        form = AddDeviceForm(initial=initial_dict)
    return render(request,
                  'device/equipment_add.html',
                  {'form': form, 'menu': menu,
                   'draft': draft,
                   'title': 'Изменение черновика'})


def draft_equipment_add(request, pk):
    if not request.user.is_staff:
        redirect('login')
    draft = Draft.objects.get(pk=pk)
    initial_dict = {
        'serial_number': draft.serial_number_draft,
        'position': draft.poz_draft,
        'model': draft.model_draft,
        'name': draft.name_draft,
        'manufacturer': draft.manufacturer_draft,
        'location': draft.location_draft,
        'tag': draft.tag_draft,
        'description': draft.description_draft,
        'status': draft.status_draft,
        'year': draft.year_draft,

    }

    if request.method == 'POST':
        form = AddEquipmentForm(request.POST)
        if form.is_valid():
            form.save(request.user)
            return redirect(reverse_lazy('draft_delete', kwargs={'pk': pk}))
    else:
        form = AddEquipmentForm(initial=initial_dict)
    return render(request,
                  'device/equipment_add.html',
                  {'form': form, 'menu': menu,
                   'draft': draft,
                   'title': 'Изменение черновика'})


def draft_delete(request, pk):
    obj = get_object_or_404(Draft, pk=pk)
    st = ('media/' + str(obj.images))
    obj.delete()
    try:
        os.remove(os.path.join(BASE_DIR, st))
        return redirect(reverse_lazy('draft_list'))
    except:
        return redirect(reverse_lazy('draft_list'))


def my_exams(request):
    if not request.user.is_authenticated:
        redirect('login')
    try:
        object = MyExam.objects.get(user=request.user)
    except:
        object = {}
    objects = MyExam.objects.all()
    try:
        initial_dict = {
            'exams_ot': object.exams_ot,
            'exams_eb': object.exams_eb,
        }
    except:
        initial_dict = {}
    if request.method == 'POST':
        form = MyExamsForm(request.POST)
        if form.is_valid():
            form.save(request.user)
            # return render(request, 'device/my_exams.html', context=context)
    else:
        form = MyExamsForm(initial=initial_dict)
    now_date = datetime.date.today() - relativedelta(months=10)
    context = {
        'form': form,
        'objects': objects,
        'object': object,
        'now_date': now_date,
    }
    return render(request, 'device/my_exams.html', context=context)


def changes(request):
    with open('./changes.csv', encoding='utf-8') as f:
        reader = DictReader(f, delimiter=';')
        count = 0
        with open('./bag.csv', 'a', encoding='utf-8') as bag:
            fieldnames = ['name', 'serial_number', 'verification']
            writer = csv.DictWriter(bag, fieldnames=fieldnames, delimiter=';')
            writer.writeheader()
            for i, row in enumerate(reader):
                try:
                    eq = Equipment.objects.get(serial_number=row['serial_number'].strip())
                except:
                    pass
                    try:
                        eq = Equipment.objects.get(serial_number=('0' + row['serial_number']))
                    except:
                        count += 1
                        writer.writerow({
                            # 'reg_number': row['reg_number'],
                            'name': row['name'],
                            'serial_number': row['serial_number'],
                            'verification': row['verification'],
                            # 'certificate': row['certificate'],
                        })
                        continue
                for i in eq.si.all():
                    previous_verification = datetime.date.fromisoformat(row['verification'])
                    i.next_verification = previous_verification + relativedelta(months=+(int(i.interval.name)))
                    # i.certificate = row['certificate']
                    # RegNumber.objects.get_or_create(name=row['reg_number'])
                    # reg = RegNumber.objects.get(name=row['reg_number'])
                    # i.reg_number = reg
                    i.save()

        sm = EmailMessage
        subject = 'sample'
        body = 'sample'
        from_email = 'freemail_2019@mail.ru'
        if not request.user.email:
            return redirect(reverse_lazy('add_email'))
        to_email = request.user.email
        msg = sm(subject, body, from_email, [to_email])
        msg.attach_file(f'{BASE_DIR}/bag.csv')
        msg.send()
    return redirect('/')


def send_bid(request, pk):
    if not request.user.is_staff:
        redirect('login')
    if not request.user.email:
        return redirect(reverse_lazy('defectone:add_email'))
    de = Equipment.objects.get(pk=pk)
    poz = GP.objects.get(name=de.positions.all().last())
    wb = load_workbook(f'{BASE_DIR}/files/bid_files/bid.xlsx')
    ws = wb['z']
    ws['C9'] = request.user.last_name
    ws['C11'] = f'{de.name.name}, {poz.construction}, (поз.{poz.name})'
    ws['C14'] = f'{de.descriptions.all().last()}'
    ws['C12'] = f'{de.name.name} {de.model}, зав.№{de.serial_number} - 1 шт., ({de.year} г.в.)'
    ws['B17'] = f'Заказчик:__________{request.user.first_name}'
    wb.save(
        f'{BASE_DIR}/files/bid_files/Заявка в рем. цех {de.name.name} {de.serial_number} от {datetime.date.today()}.xlsx')
    wb.close()
    sm = EmailMessage
    subject = 'bid'
    body = 'Заявка отправлена на почту.'
    from_email = 'freemail_2019@mail.ru'
    to_email = request.user.email
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(
        f'{BASE_DIR}/files/bid_files/Заявка в рем. цех {de.name.name} {de.serial_number} от {datetime.date.today()}.xlsx')
    msg.send()
    return redirect(reverse_lazy('search'))


class PprDateCreate(CreateView):
    permission_classes = [IsAdminUser, ]
    model = PprDate
    fields = '__all__'
    template_name = 'device/ppr_date_create.html'
    success_url = reverse_lazy('ppr_date_list')
    extra_context = {
        'menu': menu
    }


class PprDateList(ListView):
    permission_classes = [IsAdminUser, ]
    model = PprDate
    context_object_name = 'dates'
    template_name = 'device/ppr_date_list.html'
    extra_context = {
        'menu': menu
    }


class PprDateUpdate(UpdateView):
    permission_classes = [IsAdminUser, ]
    model = PprDate
    fields = '__all__'
    template_name = 'device/ppr_date_create.html'
    success_url = reverse_lazy('ppr_date_list')
    extra_context = {
        'menu': menu
    }


def get_ppr_list(request, pk):
    ppr_date = get_object_or_404(PprDate, pk=pk)
    ppr_plan = PprPlan.objects.filter(ppr=ppr_date)
    context = {
        'ppr_date': ppr_date,
        'ppr_plan': ppr_plan,
        'menu': menu
    }
    return render(request, 'device/ppr_list.html', context=context)


def ppr_create(request, pk):
    if not request.user.is_authenticated:
        redirect('login')
    ppr_date = PprDate.objects.get(pk=pk)
    ppr_plan = PprPlan.objects.filter(ppr=ppr_date)
    initial_dict = {
        'ppr': ppr_date,
    }
    if request.method == 'POST':
        form = PprForm(request.POST)
        if form.is_valid():
            form.save(request.user, pk)
            return redirect(reverse_lazy(f'ppr_date_list'))
    else:
        form = PprForm(initial=initial_dict)
    return render(request,
                  'device/ppr_add.html',
                  {'form': form, 'menu': menu,
                   'ppr_date': ppr_date,
                   'title': 'ППР'})


#
class PprUpdate(UpdateView):
    permission_classes = [IsAuthenticated]
    model = PprPlan
    fields = ['ppr', 'name', 'description', 'required_materials']
    template_name = 'device/ppr_add.html'
    success_url = reverse_lazy('ppr_list')

    # context_object_name = 'object'
    # def get_object(self, queryset=None):
    #     return get_object_or_404(Ppr, pk=self.kwargs.get('pk'))

    def form_valid(self, form):
        instance = form.save(commit=False)
        instance.user = self.request.user
        instance.save()
        return redirect('ppr_date_list')


# <a href="{% url 'ppr_list' %}" > ППР </a>

class ManualsView(ListView):
    model = Category
    context_object_name = 'categories'
    template_name = 'device/manuals.html'
    extra_context = {
        'menu': menu,
    }

# <!--<a href="{% url  'manuals' %}"> Руководства </a>-->