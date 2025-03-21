import csv
import os
from base64 import decode
from csv import excel
from datetime import datetime
from traceback import print_tb

from django.core.mail import EmailMessage
from django.shortcuts import redirect
from django.urls import reverse_lazy
from pygments.lexer import default
from openpyxl import load_workbook
from defectone.models import Defect
from device.models import Equipment, Si
from equipment.settings import BASE_DIR


def sending(request, title):
    if not request.user.is_staff:
        redirect('login')
    sm = EmailMessage
    subject = 'sample'
    body = 'sample'
    from_email = 'freemail_2019@mail.ru'
    if not request.user.email:
        return redirect(reverse_lazy('defectone:add_email'))
    to_email = request.user.email
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(f'{BASE_DIR}/from sending.xlsx')
    msg.send()
    os.remove(f'{BASE_DIR}/from sending.xlsx')
    return redirect(reverse_lazy('search'))


def sample_send(request, data):
    if not request.user.is_staff:
        redirect('login')
    if not request.user.email:
        return redirect(reverse_lazy('defectone:add_email'))
    # eq = Equipment.objects.filter(Q(serial_number=de.serial_number) & Q(model__name=de.model)).first()
    # if not eq:
    #     return redirect(reverse_lazy('defectone:defect_list'))
    wb = load_workbook(f'{BASE_DIR}/sending.xlsx')
    ws = wb['l1']

    for i, eq in enumerate(data):
        ws[f'A{i + 2}'] = i + 1
        ws[f'B{i + 2}'] = eq.serial_number
        try:
            ws[f'C{i + 2}'] = eq.positions.last().name
        except:
            ws[f'C{i + 2}'] = '-'
        try:
            ws[f'D{i + 2}'] = eq.locations.last().name
        except:
            ws[f'D{i + 2}'] = '-'
        try:
            ws[f'E{i + 2}'] = eq.tags.last().name
        except:
            ws[f'E{i + 2}'] = '-'
        ws[f'F{i + 2}'] = eq.name.name
        ws[f'G{i + 2}'] = eq.model.name
        ws[f'H{i + 2}'] = eq.year.name
        try:
            ws[f'I{i + 2}'] = eq.descriptions.last().name
        except:
            ws[f'I{i + 2}'] = '-'
        if eq.si_or:
            try:
                from_si = Si.objects.get(equipment=eq)
            except:
                continue
            # try:
            #     ws[f'J{i + 2}'] = from_si.reg_number.name
            # except:
            #     ws[f'J{i + 2}'] = 'регистрационного номера нет'
            try:
                ws[f'K{i + 2}'] = from_si.scale.max_scale
            except:
                ws[f'K{i + 2}'] = '-'
            try:
                ws[f'L{i + 2}'] = from_si.interval.name
            except:
                ws[f'L{i + 2}'] = 'interval'
            try:
                ws[f'M{i + 2}'] = from_si.unit.name
            except:
                ws[f'M{i + 2}'] = 'ед.изм'
            try:
                ws[f'J{i + 2}'] = from_si.scale.min_scale
            except:
                ws[f'J{i + 2}'] = '-'

            # try:
            #     ws[f'N{i + 2}'] = from_si.certificate
            # except:
            #     ws[f'N{i + 2}'] = 'сертификата нет'
            try:
                ws[f'N{i + 2}'] = from_si.next_verification
            except:
                ws[f'N{i + 2}'] = 'предыдущая проверка нет'
    wb.save(f'{BASE_DIR}/from sending.xlsx')
    wb.close()
    # sm = EmailMessage
    # subject = 'Samole'
    # body = 'Выборка отправлена на почту.'
    # from_email = 'freemail_2019@mail.ru'
    # to_email = request.user.email
    # msg = sm(subject, body, from_email, [to_email])
    # msg.attach_file(f'{BASE_DIR}/from sending.xlsx')
    # msg.send()


    return redirect(reverse_lazy('search'))


# def send_all(request, start, end):
#     if start == 0:
#         with open('./all_data.csv', 'w', encoding='utf-8'):
#             pass
#     if not request.user.is_staff:
#         redirect('login')
#     last = Equipment.objects.all().count()
#     get_all = Equipment.objects.filter(si_or=True)[start:end]
#     with open('./all_data.csv', 'a', encoding='utf-8') as f:
#         fieldnames = ['№', 'position', 'location', 'teg', 'model', 'name', 'serial_number', 'description',
#                       'min_scale', 'max_scale', 'comment', 'interval', 'previous_verification',
#                       'next_verification', 'result', ]  #
#         writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=';')
#         writer.writeheader()
#         for i, eq in enumerate(get_all):
#             try:
#                 from_si = Si.objects.get(equipment=eq)
#             except:
#                 continue
#             try:
#                 writer.writerow({
#                     '№': i + start,
#                     'position': eq.positions.last().name,
#                     'location': eq.locations.last().name,
#                     'description': eq.descriptions.last().name,
#                     'teg': eq.tags.last().name,
#                     # 'type': eq.type.name,
#                     'model': eq.model.name,
#                     'name': eq.name.name,
#                     # 'reg_number': from_si.reg_number,
#                     'serial_number': eq.serial_number,
#                     'min_scale': from_si.scale.min_scale,
#                     'max_scale': from_si.scale.max_scale,
#                     # 'unit': from_si.unit,
#                     'comment': eq.comment,
#                     'interval': from_si.interval,
#                     'previous_verification': from_si.previous_verification,
#                     'next_verification': from_si.next_verification,
#                     'result': from_si.result,
#                 }
#                 )
#             except:
#                 pass
#     if end < last:
#         start = end
#         end += 1000
#         return redirect(reverse_lazy('send_all', kwargs={'start': start, 'end': end}))
#     sm = EmailMessage
#     subject = 'all'
#     body = 'Выборка отправлена на почту.'
#     from_email = 'freemail_2019@mail.ru'
#     to_email = request.user.email
#     msg = sm(subject, body, from_email, [to_email])
#     msg.attach_file('./all_data.csv')
#     msg.send()
#     return redirect(reverse_lazy('search'))
#

def send_all(request, start, end):
    if not request.user.is_staff:
        redirect('login')
    last = Equipment.objects.all().count()
    get_all = Equipment.objects.filter(si_or=True)[start:end]
    if start == 0:
        wb = load_workbook(f'{BASE_DIR}/all_data.xlsx')
        ws = wb['Sheet1']
    else:
        wb = load_workbook(f'{BASE_DIR}/all.xlsx')
        ws = wb['Sheet1']
    for i, eq in enumerate(get_all):
        row = i + start + 2
        try:
            from_si = Si.objects.get(equipment=eq)
        except:
            continue
        ws[f'A{row}'] = i + start
        try:
            ws[f'B{row}'] = eq.positions.last().name
        except:
            ws[f'B{row}'] = '-'
        try:
            ws[f'C{row}'] = eq.locations.last().name
        except:
            ws[f'C{row}'] = '-'
        try:
            ws[f'D{row}'] = eq.tags.last().name
        except:
            ws[f'D{row}'] = '-'
        try:
            ws[f'H{row}'] = eq.descriptions.last().name
        except:
            ws[f'H{row}'] = '-'

        ws[f'E{row}'] = eq.model.name
        ws[f'F{row}'] = eq.name.name
        ws[f'G{row}'] = eq.serial_number
        ws[f'I{row}'] = from_si.scale.min_scale
        ws[f'J{row}'] = from_si.scale.max_scale
        ws[f'K{row}'] = from_si.unit.name
        try:
            ws[f'L{row}'] = eq.comment
        except:
            ws[f'L{row}'] = '-'
        ws[f'M{row}'] = from_si.interval.name
        ws[f'N{row}'] = from_si.previous_verification
        ws[f'O{row}'] = from_si.next_verification
        ws[f'P{row}'] = from_si.result
    wb.save(f'{BASE_DIR}/all.xlsx')
    wb.close()
    if end < last:
        start = end
        end += 1000
        return redirect(reverse_lazy('send_all', kwargs={'start': start, 'end': end}))
    sm = EmailMessage
    subject = 'all'
    body = 'Выборка отправлена на почту.'
    from_email = 'freemail_2019@mail.ru'
    to_email = request.user.email
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(f'{BASE_DIR}/all.xlsx')
    msg.send()
    os.remove(f'{BASE_DIR}/all.xlsx')
    return redirect(reverse_lazy('search'))
