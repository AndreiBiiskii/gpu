import csv
import re
from datetime import datetime
from time import sleep

from django.core.mail import EmailMessage
from django.db.models import Q
from django.shortcuts import redirect
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver import Keys
from selenium.webdriver.common.by import By
from equipment.settings import BASE_DIR

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service


def get_sample(table_tr):
    low_date = datetime.strptime('01-01-2000', '%d-%m-%Y').date()
    sample_data = ('error',)
    for i in table_tr:
        type_td = i.find_elements(By.TAG_NAME, 'td')
        try:
            if low_date < datetime.strptime('-'.join(type_td[6].text.split('.')), '%d-%m-%Y').date():
                low_date = datetime.strptime('-'.join(type_td[6].text.split('.')), '%d-%m-%Y').date()
                sample_data = ()
                for j in type_td:
                    sample_data += (j.text,)
        except:
            pass

    sleep(1)
    return sample_data


def data_from_parser(request):
    # /home/andrei/Desktop/pro/.venv/lib/python3.10/site-packages
    # options = webdriver.ChromeOptions()

    chrome_options = Options()
    # chrome_options.add_argument(r"--user-data-dir=/home/andrei/Desktop/pro/.venv/lib/python3.10/site-packages")
    chrome_options.add_argument(r"--user-data-dir=/home/user/gpu/env/lib/python3.10/site-packages")
    # chrome_options.add_argument("--headless")
    # chrome_options.add_argument("--no-sandbox")
    # chrome_options.add_argument("--disable-gpu")

    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()),
        options=chrome_options
    )
    # Version 133.0.6943.126 (Official Build) (64-bit)
    # driver = webdriver.Chrome()
    driver.get("https://fgis.gost.ru/fundmetrology/cm/results?rows=100&activeYear=%D0%92%D1%81%D0%B5")
    data_ = driver.find_element(By.CLASS_NAME, 'modal-footer')
    button = data_.find_element(By.TAG_NAME, 'button')
    button.send_keys(Keys.ENTER)
    wb_bag = load_workbook(f'{BASE_DIR}/bag.xlsx')
    ws_bag = wb_bag['bag']
    wb = load_workbook(f'{BASE_DIR}/from sending.xlsx')
    ws = wb['l1']
    wb_schema = load_workbook(f'{BASE_DIR}/schema.xlsm')
    ws_schema = wb_schema['Лист1']
    # print(ws_schema.columns)
    count = 0
    bag_count = 0
    try:
        for i, eq in enumerate(ws):
            button = driver.find_elements(By.CLASS_NAME, 'btn')
            button[1].send_keys(Keys.ENTER)
            sleep(1)
            name = driver.find_element(By.ID, 'filter_mi_mititle')
            serial_number = driver.find_element(By.ID, 'filter_mi_number')
            sleep(1)
            serial_number.clear()
            sleep(1)
            serial_number.send_keys(ws[f'B{i + 2}'].value)
            sleep(1)
            name.clear()
            sleep(1)
            name.send_keys(ws[f'F{i + 2}'].value[0:4])
            sleep(1)
            btn = driver.find_elements(By.CLASS_NAME, 'btn-primary')
            btn[0].send_keys(Keys.ENTER)
            sleep(10)
            table_div = driver.find_element(By.CLASS_NAME, 'sticky-spinner-wrap')
            table_tr = table_div.find_elements(By.TAG_NAME, 'tr')
            if len(table_tr) == 1:
                ws_bag[f'A{bag_count + 1}'] = ws[f'B{i + 2}'].value
                ws_bag[f'B{bag_count + 1}'] = ws[f'F{i + 2}'].value
                bag_count += 1
            if len(table_tr) > 1:
                for con, row in enumerate(table_tr):
                    if con == 0:
                        continue
                    table_td = row.find_elements(By.TAG_NAME, 'td')
                    if table_td[5].text != ws[f'B{i + 2}'].value:
                        continue
                    ws_schema[f'J{i + 2}'] = ws[f'H{i + 2}'].value
                    ws_schema[f'AP{i + 2}'] = ws[f'L{i + 2}'].value
                    ws_schema[f'AK{i + 2}'] = ws[f'L{i + 2}'].value
                    ws_schema[f'AU{i + 2}'] = ws[f'L{i + 2}'].value
                    ws_schema[f'L{i + 2}'] = ws[f'J{i + 2}'].value
                    ws_schema[f'O{i + 2}'] = ws[f'K{i + 2}'].value
                    ws_schema[f'N{i + 2}'] = ws[f'M{i + 2}'].value
                    ws_schema[f'Q{i + 2}'] = ws[f'M{i + 2}'].value
                    ws_schema[f'A{i + 2}'] = table_td[3].text + '  ' + table_td[4].text
                    ws_schema[f'B{i + 2}'] = table_td[2].text
                    ws_schema[f'C{i + 2}'] = table_td[1].text
                    ws_schema[f'I{i + 2}'] = table_td[5].text
                    ws_schema[f'AR{i + 2}'] = table_td[6].text
                    ws_schema[f'AS{i + 2}'] = table_td[6].text
                    ws_schema[f'AY{i + 2}'] = table_td[8].text
                    ws_schema[f'AG{i + 2}'] = table_td[0].text
                    break
            print(
                f'Проверено: {i} из {ws.max_row - 1}. Осталось: {ws.max_row - i - 1}/ Процент выполнения: {(i * 100) / (ws.max_row - 1)}%')

            count += 1
            wb_schema.save(f'{BASE_DIR}/from_schema.xlsx')

    finally:
        wb_schema.save(f'{BASE_DIR}/from_schema.xlsx')
        wb_schema.close()
        wb_bag.save(f'{BASE_DIR}/from_bag.xlsx')
        wb_bag.close()
        sm = EmailMessage
        subject = 'Schema'
        body = 'Выборка отправлена на почту.'
        from_email = 'freemail_2019@mail.ru'
        to_email = request.user.email
        msg = sm(subject, body, from_email, [to_email])
        msg.attach_file(f'{BASE_DIR}/from_schema.xlsx')
        msg.send()
        msg.attach_file(f'{BASE_DIR}/from_bag.xlsx')
        msg.send()
    wb_schema.save(f'{BASE_DIR}/from_schema.xlsx')
    wb_schema.close()
    wb_bag.save(f'{BASE_DIR}/from_bag.xlsx')
    wb_bag.close()
    sm = EmailMessage
    subject = 'Schema'
    body = 'Выборка отправлена на почту.'
    from_email = 'freemail_2019@mail.ru'
    to_email = request.user.email
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(f'{BASE_DIR}/from_schema.xlsx')
    msg.send()
    msg.attach_file(f'{BASE_DIR}/from_bag.xlsx')
    msg.send()
    return redirect('/')
