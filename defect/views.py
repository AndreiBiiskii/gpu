import random
from datetime import datetime, timedelta
from django import forms
import django_filters
from django.core.mail import send_mail, EmailMessage
from django.shortcuts import render, redirect
from django.template.context_processors import request
from django.views.generic import CreateView, ListView, DeleteView, UpdateView
from rest_framework.reverse import reverse_lazy
from openpyxl import load_workbook
from defect.models import Defect, Approve, Contractor, Kait, Worker
from device.models import Equipment, Position, Location, Status, Description, GP, Tag
from device.views import menu
from equipment.settings import BASE_DIR


def defect_act(request, poz, n):
    da = f'КАиТ-{poz}-{datetime.now().date()}-{n}'
    return da


def send_act(request, pk):
    de = Defect.objects.get(pk=pk)
    eq = Equipment.objects.get(serial_number=de.serial_number)
    wb = load_workbook(f'{BASE_DIR}/act.xlsx')
    ws = wb['act']
    ws['E9'] = de.defect_act
    ws['J5'] = de.approve.name
    ws['B5'] = de.gp
    ws['D12'] = f'{eq.name}, {eq.type}, {eq.model}'
    ws['D18'] = de.serial_number
    ws['D19'] = eq.manufacturer.name
    ws['D21'] = de.project
    ws['D23'] = de.location
    ws['D25'] = datetime.date(datetime.now())
    ws['D26'] = de.short_description
    ws['D30'] = de.causes
    ws['D32'] = de.fix
    ws['A36'] = de.contractor.name
    ws['A40'] = de.kait.name
    ws['A43'] = de.worker.name
    wb.save(f'{BASE_DIR}/act1.xlsx')

    wb.close()
    sm = EmailMessage
    subject = 'Worker'
    body = 'Дефектный акт был отправлен на почту.'
    from_email = request.user.email
    to_email = 'freemail_2019@mail.ru'
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(f'{BASE_DIR}/act1.xlsx')
    msg.send()
    return redirect(reverse_lazy('defect:defect_list'))


class DefectAdd(CreateView):
    model = Defect
    fields = ('serial_number', 'gp', 'location', 'tag', 'defect_act', 'project', 'short_description', 'causes',
              'status', 'fix', 'operating_time', 'invest_letter', 'approve', 'contractor', 'kait', 'worker',)
    success_url = '/'
    template_name = 'defect/defect_add.html'
    extra_context = {
        'title': 'Добавление дефекта',
        'menu': menu,
    }

    def get_initial(self):
        initial = super().get_initial()
        eq = Equipment.objects.get(pk=self.kwargs.get('pk'))
        initial['serial_number'] = eq.serial_number
        poz = GP.objects.get(name=Position.objects.filter(equipment=eq).last().name)
        initial['gp'] = f'{poz.name},{poz.construction}'
        initial['location'] = Location.objects.filter(equipment=eq).last().name
        initial['status'] = Status.objects.filter(equipment=eq).last().name
        initial['short_description'] = Description.objects.filter(equipment=eq).last().name
        initial['defect_act'] = defect_act(request, poz.name, eq.pk)
        initial['tag'] = Tag.objects.filter(equipment=eq).last().name
        return initial


class MyFilter(django_filters.FilterSet):
    serial_number = django_filters.CharFilter(lookup_expr='icontains', label='Серийный номер')

    gp = django_filters.CharFilter(field_name='gp',
                                   lookup_expr='exact', label='Позиция:', )
    tag = django_filters.CharFilter(field_name='tag',
                                    lookup_expr='icontains', label='Тэг:', )

    at_date = django_filters.DateFilter(
        widget=forms.TextInput(attrs=
        {
            'type': 'date',
        }), label='Дата внеесния',
        field_name='at_date', lookup_expr='icontains')

    class Meta:
        model = Defect
        fields = ['serial_number', ]


def get_last_id_from_excel(filename, sheetname="Sheet1", id_col="C"):
    wb = load_workbook('../act.xlsx')
    ws = wb[sheetname]
    ws['J5'] = 'value'
    wb.close()
    return ws


def defect_list(request):
    if not request.user.is_authenticated:
        redirect('/')
    if request.method == 'POST' and request.user.is_staff:
        eq_defect = MyFilter(request.POST, queryset=Defect.objects.all())
        data = {
            'title': 'Поиск',
            'menu': menu,
            'objects': eq_defect,
            'count': eq_defect.qs.count(),

        }
        return render(request, 'defect/defect_list.html', context=data)
    if request.method == 'GET' and request.user.is_staff:
        eq_filter = MyFilter(request.POST,
                             queryset=Defect.objects.all())
        data = {
            'title': 'Поиск',
            'menu': menu,
            'objects': eq_filter,
            'count': eq_filter.qs.count(),
        }
        return render(request, 'defect/defect_list.html', context=data)


class DefectUpdate(UpdateView):
    model = Defect
    fields = ('serial_number', 'gp', 'location', 'tag', 'defect_act', 'project', 'short_description', 'causes',
              'status', 'fix', 'operating_time', 'invest_letter', 'approve', 'contractor', 'kait', 'worker',)
    success_url = reverse_lazy('defect:defect_list')
    template_name = 'defect/defect_update.html'
    extra_context = {
        'title': 'Изменить данные',
        'menu': menu,
        # 'add': 'defect:approve_add'
    }


class ApproveAdd(CreateView):
    model = Approve
    fields = '__all__'
    success_url = reverse_lazy('defect:approves')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить подписанта',
        'menu': menu,
    }


class ApproveList(ListView):
    model = Approve
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Добавить подписанта',
        'menu': menu,
        'add': 'defect:approve_add',
        'add_delete': 'defect:approve_delete',
        'add_update': 'defect:approve_update'
    }


class ApproveDelete(DeleteView):
    model = Approve
    success_url = reverse_lazy('defect:approves')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить подписанта',
        'menu': menu,
    }


class ApproveUpdate(UpdateView):
    model = Approve
    fields = '__all__'
    success_url = reverse_lazy('defect:approves')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Изменить подписанта',
        'menu': menu,
        'add': 'defect:approve_add'
    }


class ContractorAdd(CreateView):
    model = Contractor
    fields = '__all__'
    success_url = reverse_lazy('defect:contractors')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить подрядчика',
        'menu': menu,
    }


class ContractorList(ListView):
    model = Contractor
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Добавить подрядчика',
        'menu': menu,
        'add': 'defect:contractor_add',
        'add_delete': 'defect:contractor_delete',
        'add_update': 'defect:contractor_update'
    }


class ContractorDelete(DeleteView):
    model = Contractor
    success_url = reverse_lazy('defect:contractors')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить подрядчика',
        'menu': menu,
    }


class ContractorUpdate(UpdateView):
    model = Contractor
    fields = '__all__'
    success_url = reverse_lazy('defect:contractors')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Измененить подрядчика',
        'menu': menu,
        'add': 'defect:contractor_add'
    }


class KaitAdd(CreateView):
    model = Kait
    fields = '__all__'
    success_url = reverse_lazy('defect:kaits')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить мастера по КАиТ',
        'menu': menu,
    }


class KaitList(ListView):
    model = Kait
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Добавить мастера по КАиТ',
        'menu': menu,
        'add': 'defect:kait_add',
        'add_delete': 'defect:kait_delete',
        'add_update': 'defect:kait_update'
    }



class KaitDelete(DeleteView):
    model = Kait
    success_url = reverse_lazy('defect:kaits')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить мастера по КАиТ',
        'menu': menu,
    }


class KaitUpdate(UpdateView):
    model = Kait
    fields = '__all__'
    success_url = reverse_lazy('defect:approves')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Изменить мастера по КАиТ',
        'menu': menu,
        'add': 'defect:kait_add'
    }


class WorkerAdd(CreateView):
    model = Worker
    fields = '__all__'
    success_url = reverse_lazy('defect:workers')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить мастера по цеху',
        'menu': menu,
    }
    success_message = 'Сотрудник успешно добавлен'


class WorkerList(ListView):
    model = Worker
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Мастера цеха',
        'menu': menu,
        'add': 'defect:worker_add',
        'add_delete': 'defect:worker_delete',
        'add_update': 'defect:worker_update'
    }


class WorkerDelete(DeleteView):
    model = Worker
    success_url = reverse_lazy('defect:workers')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить мастера по цеху',
        'menu': menu,
    }


class WorkerUpdate(UpdateView):
    model = Worker
    fields = '__all__'
    success_url = reverse_lazy('defect:workers')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Изменить мастера по цеху',
        'menu': menu,
        'add': 'defect:worker_add'
    }
