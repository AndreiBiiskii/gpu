from datetime import datetime
from time import sleep

from django import forms
import django_filters
from django.contrib.auth.models import User
from django.core.mail import EmailMessage
from django.db.models import Q
from django.http import HttpRequest
from django.shortcuts import render, redirect
from django.template.context_processors import request
from django.views.generic import CreateView, ListView, DeleteView, UpdateView
from rest_framework.permissions import IsAdminUser
from rest_framework.reverse import reverse_lazy
from openpyxl import load_workbook

from defectone.forms import AddUserForm, DefectAddForm
from defectone.models import Defect, Approve, Contractor, Kait, Worker
from device.models import Equipment, Position, Location, Status, Description, GP, Tag, Si
from device.views import menu
from equipment.settings import BASE_DIR


def defect_act(request, poz, n):
    da = f'АКТ КАиТ-{poz}-{datetime.date(datetime.now()).strftime("%d-%m-%Y")}-{n}'
    return da


def send_act(request, pk):
    if not request.user.is_staff:
        redirect('login')
    if not request.user.email:
        return redirect(reverse_lazy('defectone:add_email'))
    de = Defect.objects.get(pk=pk)
    eq = Equipment.objects.filter(Q(serial_number=de.serial_number) & Q(model__name=de.model)).first()
    if not eq:
        return redirect(reverse_lazy('defectone:defect_list'))
    wb = load_workbook(f'{BASE_DIR}/files/defect_files/act1.xlsx')
    ws = wb['act']
    ws['E9'] = de.defect_act
    ws['J5'] = de.approve.name
    ws['B5'] = de.gp
    ws['H2'] = de.approve.job_title
    ws['D12'] = f'{eq.name}, {eq.type}, {eq.model}'
    ws['D18'] = de.serial_number
    ws['D19'] = eq.manufacturer.name
    ws['D21'] = de.project
    ws['D23'] = de.location
    ws['D25'] = datetime.date(datetime.now()).strftime('%d-%m-%Y')
    ws['D26'] = de.short_description
    ws['D30'] = de.causes
    ws['D32'] = de.fix
    ws['D34'] = de.operating_time
    ws['A36'] = de.contractor.job_title
    ws['J37'] = de.contractor.name
    ws['A40'] = de.kait.job_title
    ws['J41'] = de.kait.name
    ws['A43'] = de.worker.job_title
    ws['J44'] = de.worker.name
    wb.save(f'{BASE_DIR}/files/defect_files/{de.defect_act}.xlsx')
    wb.close()
    sm = EmailMessage
    subject = 'Worker'
    body = 'Дефектный акт был отправлен на почту.'
    from_email = 'freemail_2019@mail.ru'
    to_email = request.user.email
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(f'{BASE_DIR}/files/defect_files/{de.defect_act}.xlsx')
    msg.send()
    return redirect(reverse_lazy('defectone:defect_list'))


def send_poverka(request):
    wb = load_workbook(f'{BASE_DIR}/poverka.xlsx')
    ws = wb['z']
    count = 0
    for row in ws:
        try:
            eq = Equipment.objects.get(serial_number=ws[f'G{count + 13}'].value, si_or=True)
            ws[f'C{count + 13}'] = eq.positions.last().name
            ws[f'D{count + 13}'] = eq.locations.last().name
            ws[f'E{count + 13}'] = eq.tags.last().name
            ws[f'M{count + 13}'] = Si.objects.get(equipment=eq).certificate
            print('value')
        except:
            print('Error')
        count += 1
    wb.save(f'{BASE_DIR}/poverka.xlsx')
    wb.close()
    sm = EmailMessage
    subject = 'Worker'
    body = 'Дефектный акт был отправлен на почту.'
    from_email = request.user.email
    to_email = 'freemail_2019@mail.ru'
    msg = sm(subject, body, from_email, [to_email])
    msg.attach_file(f'{BASE_DIR}/poverka.xlsx')
    msg.send()
    return redirect(reverse_lazy('defectone:defect_list'))


class DefectAdd(CreateView):
    form_class = DefectAddForm
    success_url = '/'
    template_name = 'defect/defect_add.html'
    extra_context = {
        'title': 'Добавление дефекта',
        'menu': menu,
    }

    def form_valid(self, form):
        instance = form.save(commit=False)
        eq = Equipment.objects.filter(
            Q(serial_number=instance.serial_number) & Q(model__name=instance.model)).first()
        eq.defect_or = True
        eq.save()
        instance.save()

        return redirect('defectone:defect_list')

    def get_initial(self):
        initial = super().get_initial()
        eq = Equipment.objects.get(pk=self.kwargs.get('pk'))
        initial['defect'] = eq
        initial['serial_number'] = eq.serial_number
        initial['model'] = eq.model.name
        initial['manufacture'] = eq.manufacturer.name
        try:
            poz = GP.objects.get(name=Position.objects.filter(equipment=eq).last().name)
            initial['gp'] = f'{poz.name},{poz.construction}'
        except:
            poz = '-'
            initial['gp'] = poz

        initial['location'] = Location.objects.filter(equipment=eq).last().name
        initial['status'] = Status.objects.filter(equipment=eq).last().name
        initial['short_description'] = Description.objects.filter(equipment=eq).last().name
        if poz == '-':
            initial['defect_act'] = defect_act(request, poz,
                                               Defect.objects.filter(gp=f'{poz.name},{poz.construction}').count() + 1)
        else:
            initial['defect_act'] = defect_act(request, poz.name,
                                               Defect.objects.filter(gp=f'{poz.name},{poz.construction}').count() + 1)
        initial['tag'] = Tag.objects.filter(equipment=eq).last().name
        return initial


class MyFilter(django_filters.FilterSet):
    serial_number = django_filters.CharFilter(lookup_expr='icontains', label='Серийный номер')
    status = django_filters.CharFilter(lookup_expr='icontains', label='Статус')
    defect_act = django_filters.CharFilter(lookup_expr='icontains', label='Дефектный акт')
    gp = django_filters.CharFilter(field_name='gp',
                                   lookup_expr='icontains', label='Позиция:', )
    tag = django_filters.CharFilter(field_name='tag',
                                    lookup_expr='icontains', label='Тэг:', )

    at_date = django_filters.DateFilter(
        widget=forms.TextInput(attrs=
        {
            'type': 'date',
        }), label='Дата внеесния',
        field_name='at_date', lookup_expr='icontains')

    class Meta:
        model = Defect
        fields = ['serial_number', 'defect_act', 'gp', 'location', 'tag', 'status', 'invest_letter', 'approve',
                  'contractor', 'kait', 'worker', 'at_date']


def defect_list(request):
    if not request.user.is_authenticated:
        redirect('/')
    if request.method == 'POST' and request.user.is_staff:
        eq_defect = MyFilter(request.POST, queryset=Defect.objects.all())
        data = {
            'title': 'Поиск',
            'menu': menu,
            'objects': eq_defect,
            'count': eq_defect.qs.count(),

        }
        return render(request, 'defect/defect_list.html', context=data)
    if request.method == 'GET' and request.user.is_staff:
        eq_filter = MyFilter(request.POST,
                             queryset=Defect.objects.all())
        data = {
            'title': 'Поиск',
            'menu': menu,
            'objects': eq_filter,
            'count': eq_filter.qs.count(),
        }
        return render(request, 'defect/defect_list.html', context=data)


class DefectUpdate(UpdateView):
    model = Defect
    fields = ('serial_number', 'gp', 'location', 'tag', 'defect_act', 'project', 'short_description', 'causes',
              'status', 'fix', 'operating_time', 'invest_letter', 'approve', 'contractor', 'kait', 'worker',)
    # success_url = reverse_lazy('defectone:defect_list')
    template_name = 'defect/defect_update.html'
    extra_context = {
        'title': 'Изменить данные',
        'menu': menu,
        # 'add': 'defectone:approve_add'
    }


class ApproveAdd(CreateView):
    model = Approve
    fields = '__all__'
    success_url = reverse_lazy('defectone:approves')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить подписанта',
        'menu': menu,
    }


class ApproveList(ListView):
    model = Approve
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Добавить подписанта',
        'menu': menu,
        'add': 'defectone:approve_add',
        'add_delete': 'defectone:approve_delete',
        'add_update': 'defectone:approve_update'
    }


class ApproveDelete(DeleteView):
    model = Approve
    success_url = reverse_lazy('defectone:approves')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить подписанта',
        'menu': menu,
    }


class ApproveUpdate(UpdateView):
    model = Approve
    fields = '__all__'
    success_url = reverse_lazy('defectone:approves')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Изменить подписанта',
        'menu': menu,
        'add': 'defectone:approve_add'
    }


class ContractorAdd(CreateView):
    model = Contractor
    fields = '__all__'
    success_url = reverse_lazy('defectone:contractors')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить подрядчика',
        'menu': menu,
    }


class ContractorList(ListView):
    model = Contractor
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Добавить подрядчика',
        'menu': menu,
        'add': 'defectone:contractor_add',
        'add_delete': 'defectone:contractor_delete',
        'add_update': 'defectone:contractor_update'
    }


class ContractorDelete(DeleteView):
    model = Contractor
    success_url = reverse_lazy('defectone:contractors')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить подрядчика',
        'menu': menu,
    }


class ContractorUpdate(UpdateView):
    model = Contractor
    fields = '__all__'
    success_url = reverse_lazy('defectone:contractors')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Измененить подрядчика',
        'menu': menu,
        'add': 'defectone:contractor_add'
    }


class KaitAdd(CreateView):
    model = Kait
    fields = '__all__'
    success_url = reverse_lazy('defectone:kaits')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить мастера по КАиТ',
        'menu': menu,
    }


class KaitList(ListView):
    model = Kait
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Добавить мастера по КАиТ',
        'menu': menu,
        'add': 'defectone:kait_add',
        'add_delete': 'defectone:kait_delete',
        'add_update': 'defectone:kait_update'
    }


class KaitDelete(DeleteView):
    model = Kait
    success_url = reverse_lazy('defectone:kaits')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить мастера по КАиТ',
        'menu': menu,
    }


class KaitUpdate(UpdateView):
    model = Kait
    fields = '__all__'
    success_url = reverse_lazy('defectone:approves')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Изменить мастера по КАиТ',
        'menu': menu,
        'add': 'defectone:kait_add'
    }


class WorkerAdd(CreateView):
    model = Worker
    fields = '__all__'
    success_url = reverse_lazy('defectone:workers')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Добавить мастера по цеху',
        'menu': menu,
    }
    success_message = 'Сотрудник успешно добавлен'


class WorkerList(ListView):
    model = Worker
    template_name = 'defect/approve_list.html'
    context_object_name = 'objects'
    extra_context = {
        'title': 'Мастера цеха',
        'menu': menu,
        'add': 'defectone:worker_add',
        'add_delete': 'defectone:worker_delete',
        'add_update': 'defectone:worker_update'
    }


class WorkerDelete(DeleteView):
    model = Worker
    success_url = reverse_lazy('defectone:workers')
    template_name = 'defect/approve_delete.html'
    context_object_name = 'object'
    extra_context = {
        'title': 'Удалить мастера по цеху',
        'menu': menu,
    }


class WorkerUpdate(UpdateView):
    model = Worker
    fields = '__all__'
    success_url = reverse_lazy('defectone:workers')
    template_name = 'defect/approve_add.html'
    extra_context = {
        'title': 'Изменить мастера по цеху',
        'menu': menu,
        'add': 'defectone:worker_add'
    }


class AddEmail(CreateView):
    model = User
    form_class = AddUserForm
    template_name = 'defect/approve_add.html'
    success_message = 'Email успешно добавлен'

    def form_valid(self, form):
        user = self.request.user
        form.save(user)
        return redirect('defectone:defect_list')


def bid(request):
    print()
    return redirect('/')
